"""Consolidação da trilha auditável em Excel e relatório Markdown."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def parecer_final(row: pd.Series) -> str:
    partes = [
        f"BERTimbau: rótulo={row.get('label','')}; predição={row.get('bert_pred','')}; tipo={row.get('tipo_erro','')}; confiança={row.get('bert_conf','')}.",
        f"ADTC: zona={row.get('zona_adtc','')}; prioridade={row.get('prioridade_rag_integrado','')}.",
        f"Contexto necessário: {row.get('tipo_contexto_necessario','')}.",
    ]
    if str(row.get("marcadores", "")).strip():
        partes.append(f"Marcadores: {row.get('marcadores')}.")
    if str(row.get("taxonomia_recursos", "")).strip():
        partes.append(f"Taxonomia: {row.get('taxonomia_recursos')}.")
    frag = str(row.get("fragmento_top1_rag_integrado", "")).strip()
    if frag:
        partes.append(
            f"RAG top1: {frag[:520]} (score integrado={row.get('score_integrado_top1_rag_integrado','')}; "
            f"score discursivo={row.get('score_discursivo_top1_rag_integrado','')}; origem={row.get('origem_contexto_top1_rag_integrado','')})."
        )
    else:
        partes.append("RAG integrado: sem fragmento top1 recuperado.")
    decisao = row.get("manual_decisao_sobre_contexto", "")
    if str(decisao).strip():
        partes.append(f"Avaliação manual: {decisao}; melhor top-k={row.get('manual_melhor_topk','')}.")
    return " ".join(partes)


def _reduzir_objetos(df: pd.DataFrame, limite: int = 3200) -> pd.DataFrame:
    out = df.copy()
    for coluna in out.select_dtypes(include="object"):
        if any(t in coluna.lower() for t in ("text", "fragmento", "consulta", "parecer")):
            out[coluna] = out[coluna].fillna("").astype(str).str.replace(r"\s+", " ", regex=True).str[:limite]
    return out


def gerar_planilha(
    caminho: str | Path,
    dataset: pd.DataFrame,
    base_rag: pd.DataFrame,
    taxonomia: pd.DataFrame,
    metricas_bert: pd.DataFrame,
    matriz_bert: pd.DataFrame,
    diagnostico_integridade: pd.DataFrame,
    diagnostico_rag: pd.DataFrame,
    amostra_manual: pd.DataFrame | None = None,
    metricas_manuais: pd.DataFrame | None = None,
    xai_sintese: pd.DataFrame | None = None,
) -> Path:
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    dados = dataset.copy()
    dados["parecer_final_adtc_rag_integrado_xai"] = dados.apply(parecer_final, axis=1)
    dados = _reduzir_objetos(dados)
    resumo = pd.DataFrame(
        [
            ["versao", "v8_modular"],
            ["objetivo", "Auditar as predições pré-computadas do BERTimbau por meio de ADTC, taxonomia, RAG integrado, avaliação manual e XAI opcional."],
            ["bertimbau_retreinado", False],
            ["total_teste", len(dados)],
            ["base_rag_integrada", len(base_rag)],
            ["avaliacao_manual_incorporada", bool(amostra_manual is not None and not amostra_manual.empty)],
        ], columns=["item", "valor"]
    )
    zonas = dados["zona_adtc"].value_counts(dropna=False).reset_index(name="n") if "zona_adtc" in dados else pd.DataFrame()
    tipos = dados["tipo_erro"].value_counts(dropna=False).reset_index(name="n") if "tipo_erro" in dados else pd.DataFrame()
    dicionario = pd.DataFrame(
        [
            ["zona_adtc", "Zona de risco calculada a partir da predição e da marcação superficial."],
            ["tipo_contexto_necessario", "Tipo de exterioridade/contexto sugerido para a auditoria do item."],
            ["score_semantico_top*_rag_integrado", "Score do recuperador semântico; relativo à consulta."],
            ["score_discursivo_top*_rag_integrado", "Heurística explícita informada pela ADTC; não equivale a condições de produção."],
            ["score_integrado_top*_rag_integrado", "Combinação declarada dos scores semântico e discursivo."],
            ["parecer_final_adtc_rag_integrado_xai", "Síntese auditável por item."],
        ], columns=["coluna", "descricao"]
    )

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        dados.to_excel(writer, sheet_name="Dataset_ADTC_RAG_XAI", index=False)
        zonas.to_excel(writer, sheet_name="Zonas_ADTC", index=False)
        tipos.to_excel(writer, sheet_name="Tipos_erro", index=False)
        metricas_bert.to_excel(writer, sheet_name="Metricas_BERT", index=False)
        matriz_bert.to_excel(writer, sheet_name="Matriz_BERT", index=False)
        diagnostico_integridade.to_excel(writer, sheet_name="Integridade", index=False)
        diagnostico_rag.to_excel(writer, sheet_name="Diagnostico_RAG", index=False)
        base_rag.to_excel(writer, sheet_name="Base_RAG_integrada", index=False)
        taxonomia.to_excel(writer, sheet_name="Taxonomia_ADTC", index=False)
        dicionario.to_excel(writer, sheet_name="Dicionario_colunas", index=False)
        if amostra_manual is not None and not amostra_manual.empty:
            amostra_manual.to_excel(writer, sheet_name="Avaliacao_manual_V8", index=False)
        if metricas_manuais is not None and not metricas_manuais.empty:
            metricas_manuais.to_excel(writer, sheet_name="Metricas_manuais_V8", index=False)
        if xai_sintese is not None and not xai_sintese.empty:
            xai_sintese.to_excel(writer, sheet_name="XAI_sintese", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, coluna in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                max_len = max(len(str(c.value or "")) for c in list(coluna)[:80])
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)
                for c in coluna:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
    return caminho


def gerar_relatorio_markdown(caminho: str | Path, dataset: pd.DataFrame, metricas_bert: pd.DataFrame, metricas_manuais: pd.DataFrame | None = None) -> Path:
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    linhas = [
        "# Relatório técnico — Pipeline V8 modular",
        "",
        "## Arquitetura",
        "BERTimbau pré-computado → zonas ADTC → taxonomia → RAG integrado → avaliação manual → XAI opcional.",
        "",
        "## Escopo",
        f"- Itens no teste: {len(dataset)}",
        f"- Casos de prioridade alta: {int((dataset.get('prioridade_rag_integrado', pd.Series(dtype=str)) == 'alta').sum())}",
        "- BERTimbau retreinado nesta execução: não",
        "",
        "## Métricas BERTimbau",
        metricas_bert.to_markdown(index=False),
    ]
    if metricas_manuais is not None and not metricas_manuais.empty:
        linhas += ["", "## Pertinência contextual manual", metricas_manuais.to_markdown(index=False)]
    linhas += [
        "",
        "## Nota metodológica",
        "O score discursivo é uma heurística explicitamente informada pela ADTC. Similaridade e score não substituem o julgamento de pertinência; a avaliação humana permanece a instância final.",
    ]
    caminho.write_text("\n".join(linhas), encoding="utf-8")
    return caminho
